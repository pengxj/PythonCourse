from random import choice, randrange
from string import digits, ascii_letters
from os import listdir, mkdir
from os.path import isdir
import sqlite3
from time import time
from openpyxl import Workbook, load_workbook

def generateRandomData():
    ''' 生成测试数据，共50个Excel文件，每个文件有5列随机字符串'''
    # 如果不存在子文件夹xlsxs，就创建
    if not isdir('xlsxs'):
        mkdir('xlsxs')
        
    # total表示记录总条数
    global total

    # 候选字符集
    characters = digits+ascii_letters

    # 生成50个Excel文件
    for i in range(50):
        xlsName = 'xlsxs\\'+str(i)+'.xlsx'
        
        # 随机数，每个xlsx文件的行数不一样
        totalLines = randrange(10**4)

        # 创建Workbook，获取第1个Worksheet
        wb = Workbook()
        ws = wb.worksheets[0]
        
        # 写入表头
        ws.append(['a', 'b', 'c', 'd', 'e'])
        # 随机数据，每行5个字段，每个字段30个字符
        for j in range(totalLines):
            line = [''.join((choice(characters) for ii in range(30))) for jj in range(5)]
            ws.append(line)
            total += 1
            
        # 保存xlsx文件
        wb.save(xlsName)

def eachXlsx(xlsxFn):
    '''针对每个xlsx文件的生成器'''
    # 打开Excel文件，获取第1个Worksheet
    wb = load_workbook(xlsxFn)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows):
        # 忽略表头
        if index == 0:
            continue
        yield tuple(map(lambda x:x.value, row))

def xlsx2sqlite():
    '''从批量Excel文件中导入数据到SQLite数据库'''
    # 获取所有xlsx文件名
    xlsxs = ('xlsxs\\'+fn for fn in listdir('xlsxs'))
    
    # 连接数据库，创建游标
    with sqlite3.connect('dataxlsx.db') as conn:
        cur = conn.cursor()
        for xlsx in xlsxs:
            #批量导入，减少提交事务的次数，可以提高速度
            sql = 'INSERT INTO fromxlsx VALUES(?,?,?,?,?)'
            cur.executemany(sql, eachXlsx(xlsx))
            conn.commit()

# 用来记录生成的随机数据的总行数
total = 0

# 生成随机数据
generateRandomData()

# 导入数据，并测试速度
start = time()
xlsx2sqlite()
delta = time()-start

print('导入用时：', delta)
print('导入速度（条/秒）：', total/delta)
